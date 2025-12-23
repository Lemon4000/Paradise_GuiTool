import os
import csv
import struct
import time
from typing import Dict, Tuple, List

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import serial
except Exception:
    serial = None

def _crc16_modbus(data: bytes) -> int:
    crc = 0xFFFF
    for b in data:
        crc ^= b
        for _ in range(8):
            if crc & 1:
                crc = (crc >> 1) ^ 0xA001
            else:
                crc >>= 1
    return crc & 0xFFFF

def _sum8(data: bytes) -> int:
    return sum(data) & 0xFF

def _read_protocol_cfg() -> Dict[str, str]:
    p = os.path.join('config', 'Protocol.csv')
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            r = csv.DictReader(f)
            rows = list(r)
            if rows:
                return rows[0]
    if openpyxl:
        x = os.path.join('config', 'params.xlsx')
        if os.path.exists(x):
            wb = openpyxl.load_workbook(x, data_only=True)
            if 'Protocol' in wb.sheetnames:
                ws = wb['Protocol']
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                vals = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
                return {str(k): str(v) for k, v in zip(headers, vals)}
    return {
        'Preamble': 'FC',
        'Checksum': 'CRC16_MODBUS',
        'Baud': '2000000',
        'Parity': 'N',
        'StopBits': '1',
        'Timeout': '1000',
        'TxStart': '!',
        'RxStart': '#',
        'TxDecimals': '0',
    }

def _read_group_mapping(group: str) -> List[Dict[str, str]]:
    csv_path = os.path.join('config', f'{group}组.csv')
    rows = []
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            r = csv.DictReader(f)
            for row in r:
                norm = {}
                for k, v in row.items():
                    kk = ('' if k is None else str(k)).strip().lstrip('\ufeff').lower().replace('_','').replace(' ','')
                    vv = '' if v is None else str(v).strip()
                    if vv.lower() in ('null','none','nan'):
                        vv = ''
                    if kk == 'key':
                        norm['Key'] = vv
                    elif kk == 'name':
                        norm['Name'] = vv
                    elif kk == 'unit':
                        norm['Unit'] = vv
                    elif kk == 'min':
                        norm['Min'] = vv
                    elif kk == 'max':
                        norm['Max'] = vv
                    elif kk == 'precision':
                        norm['Precision'] = vv
                    elif kk == 'default':
                        norm['Default'] = vv
                    elif kk == 'description':
                        norm['Description'] = vv
                if any(norm.get(x,'') for x in ('Key','Name','Unit','Min','Max','Precision','Default','Description')):
                    rows.append(norm)
            return rows
    if openpyxl:
        x = os.path.join('config', 'params.xlsx')
        if os.path.exists(x):
            wb = openpyxl.load_workbook(x, data_only=True)
            sheet_name = f'{group}组'
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_cells = next(ws.iter_rows(min_row=1, max_row=1))
                headers = [str(c.value) for c in header_cells]
                for rr in ws.iter_rows(min_row=2):
                    vals = [c.value for c in rr]
                    if all(v is None for v in vals):
                        continue
                    rows.append({headers[i]: '' if vals[i] is None else str(vals[i]) for i in range(len(headers))})
                return rows
    return []

def _fmt_value(v: float, precision: int) -> str:
    return f'{v:.{precision}f}'

def _payload_from_values(group: str, values: Dict[str, float], precision_map: Dict[str, int], start_char: str, decimals_override: str) -> str:
    parts = []
    for k in sorted(values.keys(), key=lambda x: int(x[1:])):
        if decimals_override != '':
            try:
                p = int(decimals_override)
            except Exception:
                p = precision_map.get(k, 2)
        else:
            p = precision_map.get(k, 2)
        parts.append(f'{k}:{_fmt_value(values[k], p)}')
    return start_char + 'WRITE' + ','.join(parts) + ';'

def _checksum_bytes(payload: bytes, algo: str) -> bytes:
    if algo.upper() == 'CRC16_MODBUS':
        crc = _crc16_modbus(payload)
        return struct.pack('<H', crc)
    if algo.upper() == 'SUM8':
        s = _sum8(payload)
        return bytes([s])
    return b''

def build_frame(group: str, values: Dict[str, float], cfg: Dict[str, str] | None = None) -> bytes:
    cfg = cfg or _read_protocol_cfg()
    rows = _read_group_mapping(group)
    precision_map = {}
    for r in rows:
        k = r.get('Key', '')
        p = r.get('Precision', '2')
        if k:
            try:
                precision_map[k] = int(p)
            except Exception:
                precision_map[k] = 2
    tx_start = (cfg.get('TxStart','!') or '!')[0]
    tx_dec = str(cfg.get('TxDecimals',''))
    payload = _payload_from_values(group, values, precision_map, tx_start, tx_dec).encode('ascii')
    cs_algo = cfg.get('Checksum', 'CRC16_MODBUS')
    preamble = cfg.get('Preamble', 'FC')
    pre_bytes = bytes.fromhex(preamble) if preamble else b''
    cs = _checksum_bytes(payload, cs_algo)
    return pre_bytes + payload + cs

def build_read_request(group: str, cfg: Dict[str, str] | None = None) -> bytes:
    cfg = cfg or _read_protocol_cfg()
    tx_start = (cfg.get('TxStart','!') or '!')[0]
    payload = f'{tx_start}READ:{group};'.encode('ascii')
    cs_algo = cfg.get('Checksum', 'CRC16_MODBUS')
    preamble = cfg.get('Preamble', 'FC')
    pre_bytes = bytes.fromhex(preamble) if preamble else b''
    cs = _checksum_bytes(payload, cs_algo)
    return pre_bytes + payload + cs

def _parse_payload(payload: bytes) -> Dict[str, float]:
    s = payload.decode('ascii')
    if not (s.startswith('#') and s.endswith(';')):
        return {}
    s = s[1:-1]
    out: Dict[str, float] = {}
    for part in s.split(','):
        if ':' not in part:
            continue
        k, v = part.split(':', 1)
        try:
            out[k] = float(v)
        except Exception:
            pass
    return out

def parse_frame(frame: bytes, cfg: Dict[str, str] | None = None) -> Dict[str, float]:
    cfg = cfg or _read_protocol_cfg()
    preamble = cfg.get('Preamble', 'FC')
    pre_len = len(bytes.fromhex(preamble)) if preamble else 0
    algo = cfg.get('Checksum', 'CRC16_MODBUS').upper()
    if algo == 'CRC16_MODBUS':
        cs_len = 2
    elif algo == 'SUM8':
        cs_len = 1
    else:
        cs_len = 0
    if len(frame) < pre_len + 1 + cs_len:
        return {}
    payload = frame[pre_len:len(frame)-cs_len] if cs_len else frame[pre_len:]
    cs_recv = frame[len(frame)-cs_len:] if cs_len else b''
    cs_calc = _checksum_bytes(payload, algo)
    if cs_len and cs_recv != cs_calc:
        return {}
    return _parse_payload(payload)

def _open_port(cfg: Dict[str, str], port: str):
    if serial is None:
        raise RuntimeError('pyserial 未安装')
    baud = int(cfg.get('Baud', '115200'))
    parity = cfg.get('Parity', 'N')
    stop = int(cfg.get('StopBits', '1'))
    timeout_ms = int(cfg.get('Timeout', '1000'))
    if parity == 'E':
        par = serial.PARITY_EVEN
    elif parity == 'O':
        par = serial.PARITY_ODD
    else:
        par = serial.PARITY_NONE
    ser = serial.Serial(port=port, baudrate=baud, bytesize=serial.EIGHTBITS, parity=par, stopbits=serial.STOPBITS_TWO if stop == 2 else serial.STOPBITS_ONE, timeout=timeout_ms/1000.0)
    return ser

def read_group(port: str, group: str, cfg: Dict[str, str] | None = None) -> Dict[str, float]:
    cfg = cfg or _read_protocol_cfg()
    req = build_read_request(group, cfg)
    ser = _open_port(cfg, port)
    try:
        ser.reset_input_buffer()
        ser.write(req)
        end_time = time.time() + int(cfg.get('Timeout', '1000'))/1000.0
        buf = bytearray()
        while time.time() < end_time:
            b = ser.read(1)
            if not b:
                continue
            buf += b
            if b == b';':
                algo = cfg.get('Checksum', 'CRC16_MODBUS').upper()
                cs_len = 2 if algo == 'CRC16_MODBUS' else (1 if algo == 'SUM8' else 0)
                more = ser.read(cs_len)
                buf += more
                return parse_frame(bytes(buf), cfg)
        return {}
    finally:
        ser.close()

def write_group(port: str, group: str, values: Dict[str, float], cfg: Dict[str, str] | None = None) -> bool:
    cfg = cfg or _read_protocol_cfg()
    rows = _read_group_mapping(group)
    for r in rows:
        k = r.get('Key', '')
        if not k:
            continue
        if k in values:
            try:
                mn = float(r.get('Min', '')) if r.get('Min', '') != '' else None
                mx = float(r.get('Max', '')) if r.get('Max', '') != '' else None
                v = float(values[k])
                if mn is not None and v < mn:
                    return False
                if mx is not None and v > mx:
                    return False
            except Exception:
                return False
    frame = build_frame(group, values, cfg)
    ser = _open_port(cfg, port)
    try:
        ser.reset_input_buffer()
        ser.write(frame)
        return True
    finally:
        ser.close()

def load_mapping(group: str) -> List[Dict[str, str]]:
    return _read_group_mapping(group)

if __name__ == '__main__':
    import sys
    args = sys.argv[1:]
    if not args:
        vals = {
            'A0': 14.0,'A1': 60.0,'A2': 45.0,'A3': 665.0,'A4': 1000.0,'A5': 300.0,
            'A6': 5.0,'A7': 2500.0,'A8': 50.0,'A9': 25.0,'A10': 25.0,'A11': 2.0,
            'A12': 2.0,'A13': 200.0,'A14': 10.0,'A15': 2400.0,'A16': 4450.0,'A17': 1500.0,
            'A18': 100.0,'A19': 115.0
        }
        f = build_frame('A', vals)
        print(f.hex())
        print(parse_frame(f))
    elif args[0] == 'read' and len(args) >= 3:
        port = args[1]
        group = args[2]
        print(read_group(port, group))
    elif args[0] == 'write' and len(args) >= 3:
        port = args[1]
        group = args[2]
        kvs = {}
        for kv in args[3:]:
            if '=' in kv:
                k, v = kv.split('=', 1)
                try:
                    kvs[k] = float(v)
                except Exception:
                    pass
        ok = write_group(port, group, kvs)
        print('OK' if ok else 'FAIL')
    else:
        print('用法:')
        print('  python Usart_Para_FK.py               # 构建并解析示例 A 组帧')
        print('  python Usart_Para_FK.py read COMx A   # 读取 A 组')
        print('  python Usart_Para_FK.py write COMx A A0=14.00 A1=60.00 ...')

def format_group_csv(group: str) -> bool:
    rows = _read_group_mapping(group)
    cols = ['Key','Name','Unit','Min','Max','Precision','Default','Description']
    if not rows:
        return False
    for r in rows:
        for c in cols:
            v = r.get(c, '')
            if v is None:
                v = ''
            s = str(v).strip()
            if s.lower() in ('null','none','nan'):
                s = ''
            r[c] = s
    rows.sort(key=lambda r: int(str(r.get('Key','A0'))[1:]) if str(r.get('Key','A0')).startswith('A') else 0)
    widths = {c: len(c) for c in cols}
    for r in rows:
        for c in cols:
            widths[c] = max(widths[c], len(str(r.get(c,''))))
    lines = []
    header = ",".join([str(c).ljust(widths[c]) for c in cols])
    lines.append(header)
    for r in rows:
        line = ",".join([str(r.get(c,'')).ljust(widths[c]) for c in cols])
        lines.append(line)
    out = "\n".join(lines) + "\n"
    p = os.path.join('config', f'{group}组.csv')
    with open(p, 'w', encoding='utf-8-sig', newline='') as f:
        f.write(out)
    return True
